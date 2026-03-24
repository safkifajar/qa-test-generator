import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = ["Name", "Status", "Priority", "Type", "Step", "Expected Result"]

HEADER_COLOR = "4472C4"
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_COLOR = "DCE6F1"

COL_WIDTHS = {
    "A": 45,  # Name
    "B": 10,  # Status
    "C": 10,  # Priority
    "D": 12,  # Type
    "E": 50,  # Step
    "F": 50,  # Expected Result
}


def _make_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style(cell):
    cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _make_border()


def _data_style(cell, alt_row=False):
    if alt_row:
        cell.fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = _make_border()


def export_to_excel(test_cases: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Header
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        _header_style(cell)
    ws.row_dimensions[1].height = 25

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    current_row = 2
    alt_row_toggle = False

    for tc in test_cases:
        steps = tc.get("steps", [])
        if not steps:
            continue

        num_steps = len(steps)
        start_row = current_row
        end_row = current_row + num_steps - 1
        is_alt = alt_row_toggle

        for step_idx, step_data in enumerate(steps):
            row = current_row + step_idx

            # Name — hanya baris pertama, sisanya kosong (di-merge nanti)
            name_val = tc.get("name", "") if step_idx == 0 else ""
            name_cell = ws.cell(row=row, column=1, value=name_val)
            _data_style(name_cell, is_alt)

            # Status
            status_cell = ws.cell(row=row, column=2, value="Draft" if step_idx == 0 else "")
            _data_style(status_cell, is_alt)

            # Priority
            priority_cell = ws.cell(row=row, column=3, value=tc.get("priority", "") if step_idx == 0 else "")
            _data_style(priority_cell, is_alt)

            # Type
            type_cell = ws.cell(row=row, column=4, value=tc.get("type", "") if step_idx == 0 else "")
            _data_style(type_cell, is_alt)

            # Step
            step_cell = ws.cell(row=row, column=5, value=step_data.get("step", ""))
            _data_style(step_cell, is_alt)

            # Expected Result
            result_cell = ws.cell(row=row, column=6, value=step_data.get("expected_result", ""))
            _data_style(result_cell, is_alt)

        # Merge cells kolom Name, Status, Priority, Type untuk test case ini
        if num_steps > 1:
            for col in [1, 2, 3, 4]:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row,
                    end_column=col,
                )
                merged_cell = ws.cell(row=start_row, column=col)
                merged_cell.alignment = Alignment(
                    vertical="top", wrap_text=True, horizontal="left"
                )

        current_row += num_steps
        alt_row_toggle = not alt_row_toggle

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:F1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
